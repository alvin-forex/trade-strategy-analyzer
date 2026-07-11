#!/usr/bin/env python3
"""Generate CCY × Month × Macro Analysis Excel for P10 Portfolio."""
import csv, os
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KEY_SIGNALS = ['17547','21698','22200','31593','3291','5001','14158','32278','16596','16698']
SIGNAL_EA = {s:'DW' for s in KEY_SIGNALS}

AUD_CCYS = ['AUDCAD','AUDCHF','AUDJPY','AUDNZD','AUDUSD','EURAUD','GBPAUD']
JPY_CCYS = ['AUDJPY','CADJPY','CHFJPY','EURJPY','GBPJPY','NZDJPY','USDJPY']
USD_CCYS = ['AUDUSD','EURUSD','GBPUSD','NZDUSD','USDCAD','USDCHF','USDJPY']

HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
WIN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LOSS_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

def load_all_data():
    all_data = {}
    for sid in KEY_SIGNALS:
        f = f'downloads/forex-forest-signals-page-{sid}.csv'
        if not os.path.exists(f):
            f = f'samples/{sid}/signal_{sid}_trades.csv'
        if not os.path.exists(f):
            continue
        with open(f, encoding='utf-8-sig') as fh:
            rows = list(csv.DictReader(fh))
        monthly = defaultdict(lambda: defaultdict(lambda: {'trades':0,'wins':0,'pnl':0.0,'pips':0.0,'buy_t':0,'buy_pnl':0.0,'sell_t':0,'sell_pnl':0.0}))
        for r in rows:
            symbol = r.get('Symbol','').strip()
            if not symbol: continue
            close_date = r.get('Close Time','')
            month_key = None
            for fmt in ['%d/%m/%Y %H:%M:%S','%d/%m/%Y %H:%M']:
                try:
                    dt = datetime.strptime(close_date, fmt)
                    month_key = dt.strftime('%Y-%m')
                    break
                except: pass
            if not month_key: continue
            pnl = float(r.get('Net Profit',0) or 0)
            pips = float(r.get('Net Pips',0) or 0)
            ttype = r.get('Type','').strip().lower()
            m = monthly[month_key][symbol]
            m['trades'] += 1
            if pnl > 0: m['wins'] += 1
            m['pnl'] += pnl
            m['pips'] += pips
            if 'buy' in ttype: m['buy_t'] += 1; m['buy_pnl'] += pnl
            elif 'sell' in ttype: m['sell_t'] += 1; m['sell_pnl'] += pnl
        all_data[sid] = dict(monthly)
    return all_data

def style_pnl(cell, val):
    if val > 0: cell.fill = WIN_FILL
    elif val < 0: cell.fill = LOSS_FILL

def style_wr(cell, wr):
    if wr < 50: cell.fill = LOSS_FILL
    elif wr < 70: cell.fill = WARN_FILL
    else: cell.fill = WIN_FILL

def write_ccy_tab(wb, title, subtitle, ccy_list, all_data):
    ws = wb.create_sheet(title)
    ws['A1'] = subtitle
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = "💡 月份綠=盈利/紅=虧損 | 勝率綠≥70%/黃50-70%/紅<50%"
    ws['A2'].font = Font(italic=True, color='666666')
    headers = ['Signal','CCY','方向','月份','交易數','勝率%','淨P&L($)','淨Pips','Buy側','Sell側']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center'); cell.border = THIN_BORDER
    row = 5
    for sid in sorted(all_data.keys()):
        for m in sorted(all_data[sid].keys()):
            for ccy in sorted(all_data[sid][m].keys()):
                if ccy not in ccy_list: continue
                d = all_data[sid][m][ccy]
                if d['trades'] == 0: continue
                wr = d['wins']/d['trades']*100
                d_str = f"BUY({d['buy_t']})" if d['buy_t']>d['sell_t'] else f"SELL({d['sell_t']})"
                ws.cell(row=row, column=1, value=sid)
                ws.cell(row=row, column=2, value=ccy)
                ws.cell(row=row, column=3, value=d_str)
                ws.cell(row=row, column=4, value=m)
                ws.cell(row=row, column=5, value=d['trades'])
                c7 = ws.cell(row=row, column=6, value=round(wr,1))
                c8 = ws.cell(row=row, column=7, value=round(d['pnl'],2))
                ws.cell(row=row, column=8, value=round(d['pips'],1))
                ws.cell(row=row, column=9, value=f"{d['buy_t']}/${d['buy_pnl']:.0f}")
                ws.cell(row=row, column=10, value=f"{d['sell_t']}/${d['sell_pnl']:.0f}")
                style_pnl(c8, d['pnl']); style_wr(c7, wr)
                for c in range(1,11): ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1
    for c in range(1,11): ws.column_dimensions[get_column_letter(c)].width = 15
    return ws

def build_excel(all_data, output_path):
    wb = openpyxl.Workbook()
    
    # Tab 1: 執行總綱
    ws1 = wb.active
    ws1.title = "1.執行總綱"
    ws1['A1'] = "📊 DW EA — CCY × 月份 × 宏觀經濟深度分析"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = "P10 Portfolio | 10 個 DW Signal | 2024-2026"
    ws1['A4'] = "📌 Tab 說明"
    ws1['A4'].font = Font(bold=True, size=12)
    tabs = [
        "Tab 2: 🇦🇺 AUD 相關 CCY 月度（金價上升周期影響）",
        "Tab 3: 🇯🇵 JPY 相關 CCY 月度（日圓政策影響）",
        "Tab 4: 🇺🇸 USD 相關 CCY 月度（美元強弱周期）",
        "Tab 5: 📅 宏觀周期 × AUD CCY 匯總矩陣",
        "Tab 6: 🔍 跨 Signal 同 CCY 對比（唔同 TF/MA 設定）",
        "Tab 7: 📝 宏觀分析筆記 + 季節性觀察",
    ]
    for i, t in enumerate(tabs):
        ws1[f'A{5+i}'] = t

    # Tabs 2-4
    write_ccy_tab(wb, "2.AUD相關CCY", "🇦🇺 AUD 相關 CCY 月度表現", AUD_CCYS, all_data)
    write_ccy_tab(wb, "3.JPY相關CCY", "🇯🇵 JPY 相關 CCY 月度表現", JPY_CCYS, all_data)
    write_ccy_tab(wb, "4.USD相關CCY", "🇺🇸 USD 相關 CCY 月度表現", USD_CCYS, all_data)

    # Tab 5: 宏觀周期匯總
    ws5 = wb.create_sheet("5.宏觀周期匯總")
    ws5['A1'] = "🌍 宏觀經濟周期 × AUD CCY 匯總"
    ws5['A1'].font = Font(bold=True, size=13)
    
    events = [
        ('2024-Q1\n加息尾聲','DXY 103-106','USD 強勢消退'),
        ('2024-Q3\nJPY 干預','USDJPY 161→139','JPY 反彈'),
        ('2024-Q4\n大選+金突破','XAU $2700','避險升溫'),
        ('2025-Q1\nTrump 關稅','DXY→110','貿易戰'),
        ('2025-Q2\nFed 減息','XAU $3000→$3500','AUD 受益'),
        ('2025-Q3\n暑假流動性','XAU $3500-3800','AUD 走強'),
        ('2025-Q4\n金價拉鋸','XAU $3800-4000','JPY 波動'),
        ('2026-Q1\n金突破$4000','XAU 創新高 $4140','結構轉變'),
        ('2026-Q2\n鷹派 Fed','DXY <101','USD 弱'),
    ]
    ws5['A3'] = "周期"
    ws5['B3'] = "市場特徵"
    ws5['C3'] = "影響"
    ws5['E3'] = "AUD CCY 匯總 P&L（跨所有 Signal）"
    for c in ['A3','B3','C3','E3']:
        ws5[c].font = HEADER_FONT; ws5[c].fill = HEADER_FILL
    ws5.merge_cells('E3:K3')
    
    for i, (period, market, impact) in enumerate(events):
        r = 4 + i
        ws5.cell(row=r, column=1, value=period).font = Font(bold=True)
        ws5.cell(row=r, column=2, value=market)
        ws5.cell(row=r, column=3, value=impact)
        ws5.cell(row=r, column=3).fill = WARN_FILL

    # AUD CCY period summary
    period_months_map = [
        ('2024-Q1', ['2024-01','2024-02','2024-03']),
        ('2024-Q2', ['2024-04','2024-05','2024-06']),
        ('2024-Q3', ['2024-07','2024-08','2024-09']),
        ('2024-Q4', ['2024-10','2024-11','2024-12']),
        ('2025-Q3', ['2025-07','2025-08','2025-09']),
        ('2025-Q4', ['2025-10','2025-11','2025-12']),
        ('2026-Q1', ['2026-01','2026-02','2026-03']),
        ('2026-Q2', ['2026-04','2026-05','2026-06']),
    ]
    
    ws5['E13'] = "AUD CCY × 季度 P&L 矩陣"
    ws5['E13'].font = Font(bold=True, size=12)
    
    hdr = ['周期'] + AUD_CCYS + ['合計']
    for c, h in enumerate(hdr, 5):
        cell = ws5.cell(row=14, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER
    
    for ri, (pname, pmonths) in enumerate(period_months_map):
        r = 15 + ri
        ws5.cell(row=r, column=5, value=pname).font = Font(bold=True)
        total = 0
        for ci, ccy in enumerate(AUD_CCYS, 6):
            pnl_sum = 0; trade_sum = 0
            for sid in all_data:
                for m in pmonths:
                    if m in all_data[sid] and ccy in all_data[sid][m]:
                        pnl_sum += all_data[sid][m][ccy]['pnl']
                        trade_sum += all_data[sid][m][ccy]['trades']
            cell = ws5.cell(row=r, column=ci, value=f"${pnl_sum:.0f}\n({trade_sum}T)" if trade_sum else "-")
            style_pnl(cell, pnl_sum)
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
            cell.border = THIN_BORDER
            total += pnl_sum
        tc = ws5.cell(row=r, column=6+len(AUD_CCYS), value=f"${total:.0f}")
        tc.font = Font(bold=True); style_pnl(tc, total); tc.border = THIN_BORDER
        ws5.cell(row=r, column=5).border = THIN_BORDER
    
    # Also add JPY and USD matrices
    for group_name, ccys in [('JPY', JPY_CCYS), ('USD', USD_CCYS)]:
        start_row = 15 + len(period_months_map) + 2
        ws5.cell(row=start_row, column=5, value=f"{group_name} CCY × 季度 P&L 矩陣").font = Font(bold=True, size=12)
        hdr2 = ['周期'] + ccys + ['合計']
        for c, h in enumerate(hdr2, 5):
            cell = ws5.cell(row=start_row+1, column=c, value=h)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = THIN_BORDER
        for ri, (pname, pmonths) in enumerate(period_months_map):
            r = start_row + 2 + ri
            ws5.cell(row=r, column=5, value=pname).font = Font(bold=True)
            total = 0
            for ci, ccy in enumerate(ccys, 6):
                pnl_sum = 0; trade_sum = 0
                for sid in all_data:
                    for m in pmonths:
                        if m in all_data[sid] and ccy in all_data[sid][m]:
                            pnl_sum += all_data[sid][m][ccy]['pnl']
                            trade_sum += all_data[sid][m][ccy]['trades']
                cell = ws5.cell(row=r, column=ci, value=f"${pnl_sum:.0f}\n({trade_sum}T)" if trade_sum else "-")
                style_pnl(cell, pnl_sum)
                cell.alignment = Alignment(wrap_text=True, horizontal='center')
                cell.border = THIN_BORDER
                total += pnl_sum
            tc = ws5.cell(row=r, column=6+len(ccys), value=f"${total:.0f}")
            tc.font = Font(bold=True); style_pnl(tc, total); tc.border = THIN_BORDER
            ws5.cell(row=r, column=5).border = THIN_BORDER
    
    for c in range(1,15): ws5.column_dimensions[get_column_letter(c)].width = 15

    # Tab 6: 跨 Signal 對比
    ws6 = wb.create_sheet("6.跨Signal對比")
    ws6['A1'] = "🔍 同一 CCY 喺唔同 Signal（唔同 DW 設定）嘅表現"
    ws6['A1'].font = Font(bold=True, size=13)
    all_ccys = sorted(set(ccy for sid in all_data for m in all_data[sid] for ccy in all_data[sid][m]))
    headers6 = ['CCY','Signal','總交易','總勝率%','總P&L','總Pips','最佳月($)','最差月($)','月度波幅','評級']
    for c, h in enumerate(headers6, 1):
        cell = ws6.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center'); cell.border = THIN_BORDER
    row = 4
    for ccy in all_ccys:
        for sid in sorted(all_data.keys()):
            total_pnl=0; total_t=0; total_w=0; total_pips=0; monthly_pnl={}
            for m in all_data[sid]:
                if ccy in all_data[sid][m]:
                    d = all_data[sid][m][ccy]
                    total_pnl += d['pnl']; total_t += d['trades']
                    total_w += d['wins']; total_pips += d['pips']
                    if d['trades']>0: monthly_pnl[m]=d['pnl']
            if total_t == 0: continue
            wr = total_w/total_t*100
            best_v = max(monthly_pnl.values()) if monthly_pnl else 0
            worst_v = min(monthly_pnl.values()) if monthly_pnl else 0
            vol = best_v - worst_v if monthly_pnl else 0
            rating = "⚠️高波幅" if vol>500 else ("⚡中等" if vol>200 else "✅穩定")
            if wr<50: rating += " 🔴低WR"
            ws6.cell(row=row, column=1, value=ccy)
            ws6.cell(row=row, column=2, value=sid)
            ws6.cell(row=row, column=3, value=total_t)
            c5 = ws6.cell(row=row, column=4, value=round(wr,1))
            c6 = ws6.cell(row=row, column=5, value=round(total_pnl,2))
            ws6.cell(row=row, column=6, value=round(total_pips,1))
            ws6.cell(row=row, column=7, value=round(best_v,2))
            ws6.cell(row=row, column=8, value=round(worst_v,2))
            ws6.cell(row=row, column=9, value=round(vol,2))
            ws6.cell(row=row, column=10, value=rating)
            style_pnl(c6, total_pnl); style_wr(c5, wr)
            for c in range(1,11): ws6.cell(row=row, column=c).border = THIN_BORDER
            row += 1
    for c in range(1,11): ws6.column_dimensions[get_column_letter(c)].width = 15

    # Tab 7: 宏觀分析筆記
    ws7 = wb.create_sheet("7.宏觀分析筆記")
    ws7['A1'] = "📝 宏觀經濟 × CCY 季節性分析筆記"
    ws7['A1'].font = Font(bold=True, size=14)
    
    notes = [
        ("🥇 金價與 AUD 的關係", True),
        ("澳洲係全球第二大黃金開採國。金價上升 → 澳洲貿易條件改善 → AUD 中長線支撐。", False),
        ("但 DW EA 係技術策略（MA crossover），對基本面反應有滯後。", False),
        ("", False),
        ("📊 2025-Q4 至 2026-Q2 金價瘋狂期（$3500→$4140）嘅觀察", True),
        ("• AUDUSD：BUY 主導，EA 捕捉到 AUD 走強趨勢，多數 Signal 都賺錢", False),
        ("• AUDJPY：SELL 集中（JPY 避險 vs AUD 商品），波幅大但方向正確", False),
        ("• AUDNZD：兩個商品貨幣互相拉鋸，WR 參差，波幅最大", False),
        ("• GBPAUD：表現分化，取決於英國經濟數據 vs 澳洲資源出口", False),
        ("", False),
        ("🔍 跨 Signal 設定差異嘅關鍵發現", True),
        ("同一 CCY 喺唔同 TF/MA 設定下表現可以差天共地：", False),
        ("• 短 TF（M5/M15）+ 大 MA = 更敏感，更多交易，波幅更大", False),
        ("• 長 TF（H1/H4）+ 小 MA = 更穩定，但可能錯過快速趨勢", False),
        ("• 同一 CCY 喺某啲月份可以 A Signal 賺 + B Signal 蝕", False),
        ("", False),
        ("📅 季節性模式觀察", True),
        ("• Q1（1-3月）：年初流動性逐步恢復，趨勢通常較清晰", False),
        ("• Q2（4-6月）：Fed 利率決策密集，CCY 波幅增加", False),
        ("• Q3（7-9月）：暑假流動性偏低，突發事件影響放大", False),
        ("• Q4（10-12月）：年底持倉調整，避險需求通常支撐黃金", False),
        ("", False),
        ("⚠️ 風險提示", True),
        ("• 過往表現唔代表未來 — 宏觀環境可能突變", False),
        ("• 馬丁加倉型策略喺低波幅月份表現好，但高波幅月份風險極大", False),
        ("• 建議結合 CCY Power + ATR 波幅 + 宏觀日曆 做實時風控", False),
    ]
    for i, (text, is_header) in enumerate(notes):
        cell = ws7.cell(row=2+i, column=1, value=text)
        if is_header:
            cell.font = Font(bold=True, size=12)
    ws7.column_dimensions['A'].width = 80

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")

if __name__ == '__main__':
    all_data = load_all_data()
    output = '/home/alvin/.openclaw/workspace/DW_Macro_CCY_Analysis_2026-07-11.xlsx'
    build_excel(all_data, output)
